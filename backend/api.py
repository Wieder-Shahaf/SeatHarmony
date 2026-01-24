from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
import os
import time
import io
import requests

# Load .env file early (before any other imports that might use env vars)
from dotenv import load_dotenv

# Try to load .env from project root (parent of backend/)
project_root = Path(__file__).parent.parent
env_file = project_root / ".env"
if env_file.exists():
    load_dotenv(env_file)
elif (Path(__file__).parent / ".env").exists():
    # Fallback to backend/.env
    load_dotenv(Path(__file__).parent / ".env")

from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, RootModel
import json

from .seat_harmony_task import SeatHarmonyTask, SeatHarmonyState
from .models import layout_to_dict
from .logger import get_logger, set_request_id, get_request_id

# Initialize logger for this module
logger = get_logger("api")


# =============================================================================
# GROQ API RATE LIMITER
# =============================================================================

class GroqRateLimiter:
    """
    Rate limiter for Groq API that tracks all rate limits:
    - RPM: 30 requests per minute
    - RPD: 1,000 requests per day
    - TPM: 12,000 tokens per minute
    - TPD: 100,000 tokens per day
    
    Uses safety margins to avoid hitting limits.
    """
    def __init__(
        self,
        rpm_limit: int = 30,
        rpd_limit: int = 1000,
        tpm_limit: int = 12000,
        tpd_limit: int = 100000,
        safety_margin: float = 0.1
    ):
        # Request limits
        self.rpm_limit = rpm_limit
        self.rpd_limit = rpd_limit
        
        # Token limits
        self.tpm_limit = tpm_limit
        self.tpd_limit = tpd_limit
        
        # Apply safety margin (use only 90% of limits)
        self.available_rpm = int(rpm_limit * (1 - safety_margin))  # 27 requests/min
        self.available_rpd = int(rpd_limit * (1 - safety_margin))  # 900 requests/day
        self.available_tpm = int(tpm_limit * (1 - safety_margin))  # 10,800 tokens/min
        self.available_tpd = int(tpd_limit * (1 - safety_margin))  # 90,000 tokens/day
        
        # Per-minute tracking: [(timestamp, tokens), ...]
        self.token_usage: List[Tuple[float, int]] = []
        self.request_times: List[float] = []  # Timestamps of requests
        
        # Per-day tracking: (date_string, tokens, requests)
        self.daily_tokens: int = 0
        self.daily_requests: int = 0
        self.daily_reset_date: Optional[str] = None
        
        # Track rate limit headers from last response
        self.last_rate_limit_headers: Dict[str, Optional[Any]] = {
            'remaining_requests': None,
            'remaining_tokens': None,
            'limit_requests': None,
            'limit_tokens': None,
            'reset_requests_seconds': None,  # Parsed reset time for requests
            'reset_tokens_seconds': None,     # Parsed reset time for tokens
        }
    
    def _get_date_string(self, timestamp: float) -> str:
        """Get date string (YYYY-MM-DD) for a timestamp."""
        import datetime
        return datetime.datetime.fromtimestamp(timestamp, tz=datetime.timezone.utc).strftime('%Y-%m-%d')
    
    def _reset_daily_tracking_if_needed(self, current_time: float):
        """Reset daily counters if we've crossed into a new day."""
        current_date = self._get_date_string(current_time)
        if self.daily_reset_date != current_date:
            logger.info(f"Rate limiter: Daily reset | old_date={self.daily_reset_date} new_date={current_date}")
            self.daily_tokens = 0
            self.daily_requests = 0
            self.daily_reset_date = current_date
    
    def _clean_old_entries(self, current_time: float):
        """Remove token usage and request entries older than 1 minute."""
        one_minute_ago = current_time - 60.0
        self.token_usage = [(ts, tokens) for ts, tokens in self.token_usage if ts > one_minute_ago]
        self.request_times = [ts for ts in self.request_times if ts > one_minute_ago]
    
    def get_available_tokens_per_minute(self) -> int:
        """Get available tokens in the current minute window."""
        current_time = time.time()
        self._clean_old_entries(current_time)
        
        used_tokens = sum(tokens for _, tokens in self.token_usage)
        available = max(0, self.available_tpm - used_tokens)
        return available
    
    def get_available_tokens_per_day(self) -> int:
        """Get available tokens in the current day."""
        current_time = time.time()
        self._reset_daily_tracking_if_needed(current_time)
        return max(0, self.available_tpd - self.daily_tokens)
    
    def get_available_requests_per_minute(self) -> int:
        """Get available requests in the current minute window."""
        current_time = time.time()
        self._clean_old_entries(current_time)
        return max(0, self.available_rpm - len(self.request_times))
    
    def get_available_requests_per_day(self) -> int:
        """Get available requests in the current day."""
        current_time = time.time()
        self._reset_daily_tracking_if_needed(current_time)
        return max(0, self.available_rpd - self.daily_requests)
    
    def record_request(self, tokens: int):
        """Record a request and its token usage for all rate limiting."""
        current_time = time.time()
        
        # Record per-minute usage
        self.token_usage.append((current_time, tokens))
        self.request_times.append(current_time)
        self._clean_old_entries(current_time)
        
        # Record per-day usage
        self._reset_daily_tracking_if_needed(current_time)
        self.daily_tokens += tokens
        self.daily_requests += 1
    
    def _parse_reset_time(self, reset_str: str) -> Optional[float]:
        """
        Parse reset time string from Groq headers.
        Formats: "5.045s", "1h52m19.2s", "1ms", "2m30s"
        Returns seconds as float, or None if parsing fails.
        """
        import re
        if not reset_str:
            return None
        
        try:
            # Handle milliseconds
            if reset_str.endswith('ms'):
                ms = float(reset_str[:-2])
                return ms / 1000.0
            
            # Handle seconds only: "5.045s"
            if reset_str.endswith('s') and 'h' not in reset_str and 'm' not in reset_str:
                return float(reset_str[:-1])
            
            # Handle complex formats: "1h52m19.2s" or "2m30s"
            total_seconds = 0.0
            
            # Extract hours: "1h"
            hour_match = re.search(r'(\d+(?:\.\d+)?)h', reset_str)
            if hour_match:
                total_seconds += float(hour_match.group(1)) * 3600
            
            # Extract minutes: "52m" or "2m"
            min_match = re.search(r'(\d+(?:\.\d+)?)m', reset_str)
            if min_match:
                total_seconds += float(min_match.group(1)) * 60
            
            # Extract seconds: "19.2s" or "30s"
            sec_match = re.search(r'(\d+(?:\.\d+)?)s', reset_str)
            if sec_match:
                total_seconds += float(sec_match.group(1))
            
            return total_seconds if total_seconds > 0 else None
        except (ValueError, AttributeError):
            return None
    
    def update_from_headers(self, headers: Dict[str, Any]):
        """
        Update rate limit tracking from API response headers.
        Headers: x-ratelimit-remaining-requests, x-ratelimit-remaining-tokens, etc.
        """
        # Groq uses standard rate limit headers
        remaining_requests = headers.get('x-ratelimit-remaining-requests')
        remaining_tokens = headers.get('x-ratelimit-remaining-tokens')
        limit_requests = headers.get('x-ratelimit-limit-requests')
        limit_tokens = headers.get('x-ratelimit-limit-tokens')
        
        if remaining_requests is not None:
            try:
                self.last_rate_limit_headers['remaining_requests'] = int(remaining_requests)
            except (ValueError, TypeError):
                pass
        
        if remaining_tokens is not None:
            try:
                self.last_rate_limit_headers['remaining_tokens'] = int(remaining_tokens)
            except (ValueError, TypeError):
                pass
        
        if limit_requests is not None:
            try:
                self.last_rate_limit_headers['limit_requests'] = int(limit_requests)
            except (ValueError, TypeError):
                pass
        
        if limit_tokens is not None:
            try:
                self.last_rate_limit_headers['limit_tokens'] = int(limit_tokens)
            except (ValueError, TypeError):
                pass
        
        reset_requests = headers.get('x-ratelimit-reset-requests')
        reset_tokens = headers.get('x-ratelimit-reset-tokens')
        
        # Parse reset times (these tell us exactly when limits reset)
        if reset_requests is not None:
            reset_seconds = self._parse_reset_time(str(reset_requests))
            if reset_seconds is not None:
                self.last_rate_limit_headers['reset_requests_seconds'] = reset_seconds
        
        if reset_tokens is not None:
            reset_seconds = self._parse_reset_time(str(reset_tokens))
            if reset_seconds is not None:
                self.last_rate_limit_headers['reset_tokens_seconds'] = reset_seconds
    
    def wait_if_needed(self, estimated_tokens: int) -> float:
        """
        Wait if we don't have enough capacity available (checks all limits).
        Returns the wait time in seconds.
        
        Checks in order:
        1. Requests per minute (RPM)
        2. Requests per day (RPD)
        3. Tokens per minute (TPM)
        4. Tokens per day (TPD)
        """
        current_time = time.time()
        self._clean_old_entries(current_time)
        self._reset_daily_tracking_if_needed(current_time)
        
        # Check RPM limit
        available_rpm = self.get_available_requests_per_minute()
        if available_rpm <= 0:
            # Need to wait for requests to expire
            if self.request_times:
                oldest_request = min(self.request_times)
                wait_time = 60.0 - (current_time - oldest_request) + 0.5
                wait_time = max(0, min(wait_time, 60.0))
                if wait_time > 0:
                    logger.warning(f"Rate limiter: RPM limit reached ({self.available_rpm}/{self.rpm_limit}), waiting {wait_time:.2f}s")
                    time.sleep(wait_time)
                    self._clean_old_entries(time.time())
                    return wait_time
        
        # Check RPD limit
        available_rpd = self.get_available_requests_per_day()
        if available_rpd <= 0:
            # Need to wait until next day
            wait_time = self._get_seconds_until_midnight(current_time)
            logger.error(f"Rate limiter: RPD limit reached ({self.daily_requests}/{self.rpd_limit}), must wait {wait_time:.0f}s until reset")
            if wait_time > 0:
                time.sleep(min(wait_time, 3600))  # Cap wait at 1 hour
            return wait_time
        
        # Check TPM limit
        available_tpm = self.get_available_tokens_per_minute()
        if available_tpm < estimated_tokens:
            needed_tokens = estimated_tokens - available_tpm
            
            if self.token_usage:
                # Calculate when enough tokens will be available
                expiring_entries = sorted(
                    [(ts + 60.0, tokens) for ts, tokens in self.token_usage],
                    key=lambda x: x[0]
                )
                
                tokens_freed = 0
                for expire_time, tokens in expiring_entries:
                    tokens_freed += tokens
                    if tokens_freed >= needed_tokens:
                        wait_time = expire_time - current_time + 0.5
                        wait_time = max(0, min(wait_time, 60.0))
                        if wait_time > 0:
                            logger.warning(f"Rate limiter: TPM limit reached (need {estimated_tokens}, have {available_tpm}), waiting {wait_time:.2f}s")
                            time.sleep(wait_time)
                            self._clean_old_entries(time.time())
                            return wait_time
                
                # Fallback: wait for oldest entry
                oldest_entry_time = min(ts for ts, _ in self.token_usage)
                wait_time = 60.0 - (current_time - oldest_entry_time) + 1.0
                wait_time = max(0, min(wait_time, 60.0))
                if wait_time > 0:
                    logger.warning(f"Rate limiter: TPM limit reached, waiting {wait_time:.2f}s (fallback)")
                    time.sleep(wait_time)
                    self._clean_old_entries(time.time())
                    return wait_time
        
        # Check TPD limit
        available_tpd = self.get_available_tokens_per_day()
        if available_tpd < estimated_tokens:
            wait_time = self._get_seconds_until_midnight(current_time)
            logger.error(f"Rate limiter: TPD limit reached (need {estimated_tokens}, have {available_tpd}/{self.tpd_limit}), must wait {wait_time:.0f}s until reset")
            if wait_time > 0:
                time.sleep(min(wait_time, 3600))  # Cap wait at 1 hour
            return wait_time
        
        return 0
    
    def _get_seconds_until_midnight(self, current_time: float) -> float:
        """Calculate seconds until midnight UTC."""
        import datetime
        now = datetime.datetime.fromtimestamp(current_time, tz=datetime.timezone.utc)
        midnight = (now + datetime.timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        return (midnight - now).total_seconds()
    
    def parse_rate_limit_error(self, error_str: str) -> Optional[float]:
        """
        Parse Groq rate limit error to extract wait time.
        Example: "Please try again in 2.865s"
        """
        import re
        # Look for "try again in X.Xs" pattern
        match = re.search(r'try again in ([\d.]+)s', error_str, re.IGNORECASE)
        if match:
            return float(match.group(1))
        return None
    
    def get_status(self) -> Dict[str, Any]:
        """Get current rate limiter status for debugging."""
        current_time = time.time()
        self._clean_old_entries(current_time)
        self._reset_daily_tracking_if_needed(current_time)
        
        return {
            'rpm': {
                'used': len(self.request_times),
                'available': self.get_available_requests_per_minute(),
                'limit': self.rpm_limit,
            },
            'rpd': {
                'used': self.daily_requests,
                'available': self.get_available_requests_per_day(),
                'limit': self.rpd_limit,
            },
            'tpm': {
                'used': sum(tokens for _, tokens in self.token_usage),
                'available': self.get_available_tokens_per_minute(),
                'limit': self.tpm_limit,
            },
            'tpd': {
                'used': self.daily_tokens,
                'available': self.get_available_tokens_per_day(),
                'limit': self.tpd_limit,
            },
            'headers': self.last_rate_limit_headers.copy(),
        }

# Global rate limiter instance
_groq_rate_limiter = GroqRateLimiter()


def _call_groq_with_usage(prompt: str, model: str = "llama-3.3-70b-versatile", temperature: float = 0.4, max_tokens: int = 400) -> Tuple[str, int]:
    """
    Call Groq API directly and return both response text and actual token usage.
    Also updates the rate limiter with response headers.
    Returns: (response_text, total_tokens)
    Raises: requests.HTTPError for non-2xx responses (including 429)
    """
    groq_api_key = os.getenv("GROQ_API_KEY", "")
    if not groq_api_key:
        raise ValueError("GROQ_API_KEY not found in environment")
    
    headers = {
        "Authorization": f"Bearer {groq_api_key}",
        "Content-Type": "application/json"
    }
    
    payload = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    
    response = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers=headers,
        json=payload,
        timeout=60
    )
    
    # Extract rate limit headers before raising for status (important for 429 errors)
    rate_limit_headers = {}
    for header_name in response.headers:
        if 'ratelimit' in header_name.lower():
            rate_limit_headers[header_name.lower()] = response.headers[header_name]
    
    # Update rate limiter with headers (even if request failed - headers are still useful)
    if rate_limit_headers:
        _groq_rate_limiter.update_from_headers(rate_limit_headers)
        logger.debug(f"Rate limit headers updated: {rate_limit_headers}")
    
    # Raise for status will throw HTTPError for 4xx/5xx responses
    # This allows callers to catch and handle 429 specifically
    try:
        response.raise_for_status()
    except requests.HTTPError as e:
        # Log rate limit headers if available for debugging
        if response.status_code == 429 and rate_limit_headers:
            logger.warning(f"429 Rate limit error | headers={rate_limit_headers}")
        raise
    
    data = response.json()
    
    # Extract text and usage
    text = data["choices"][0]["message"]["content"]
    usage = data.get("usage", {})
    total_tokens = usage.get("total_tokens", 0)
    
    # Fallback to estimation if usage not provided
    if total_tokens == 0:
        total_tokens = int(len(prompt) * 1.3) + int(len(text) * 1.3)
    
    return text, total_tokens


class GuestIn(BaseModel):
    id: str
    name: str
    group_id: Optional[str] = None
    importance: int = 0
    tags: List[str] = []


class TableIn(BaseModel):
    id: str
    name: str
    capacity: int
    zone: Optional[str] = None
    constraints: Dict[str, Any] = {}


class SettingsIn(RootModel[Dict[str, Any]]):
    root: Dict[str, Any] = {}


class TotParams(BaseModel):
    depth: int = 2        # 2 levels = 3 + 9 = 12 optimizations for diverse results
    branching: int = 3    # 3 children per node
    n_generate: int = 3   # Generate 3 thought variants
    n_evaluate: int = 3   # Evaluate top 3
    top_k: int = 3        # Return top 3 layouts


class LayoutRequest(BaseModel):
    guests: List[GuestIn]
    tables: List[TableIn]
    settings: Dict[str, Any] = {}
    tot: TotParams = TotParams()


class ExplainRequest(BaseModel):
    layout: Dict[str, Any]


class ExplainGuestsRequest(BaseModel):
    guests: List[GuestIn]
    tables: List[TableIn]
    layout: Dict[str, Any]  # The layout with assignments
    weights: Dict[str, float]  # The weights used for this layout
    notes: str  # The strategy/thought name (e.g., "traditional_seating")


class ExcelExportRequest(BaseModel):
    guests: List[GuestIn]
    tables: List[TableIn]
    layout: Dict[str, Any]
    options: Dict[str, bool] = {
        "include_dietary": True,
        "include_vendor_summary": False,
        "include_table_details": True,
    }


app = FastAPI(title="SeatHarmony ToT API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_event():
    """Verify API key is available on startup."""
    logger.info("=" * 60)
    logger.info("SeatHarmony API starting up...")
    logger.info("=" * 60)

    # .env is already loaded at module import time (above)
    # Just verify the Groq key is available
    groq_key = os.getenv("GROQ_API_KEY")

    if groq_key:
        logger.info("GROQ_API_KEY found in environment - using Llama 3.3 70B")
    else:
        logger.warning("GROQ_API_KEY not found. Set GROQ_API_KEY in .env file for AI features")
        logger.warning(f"Looked for .env at: {env_file} or {Path(__file__).parent / '.env'}")

    logger.info("API startup complete - ready to receive requests")


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================


def _group_guests_by_table(
    assignments: Dict[str, str],
    guests_dict: Dict[str, Any]
) -> Dict[str, List[Any]]:
    """
    Group guests by their assigned table.

    Args:
        assignments: Dict mapping guest_id -> table_id
        guests_dict: Dict mapping guest_id -> guest object/dict

    Returns:
        Dict mapping table_id -> list of guest objects/dicts
    """
    table_to_guests: Dict[str, List[Any]] = {}
    for guest_id, table_id in assignments.items():
        if table_id not in table_to_guests:
            table_to_guests[table_id] = []
        if guest_id in guests_dict:
            table_to_guests[table_id].append(guests_dict[guest_id])
    return table_to_guests


def _sequential_tot_bfs(
    instance: Dict[str, Any],
    depth: int,
    branching: int = 3,
    n_generate: int = 3,
    n_evaluate: int = 3,
) -> List[Tuple[SeatHarmonyState, float]]:
    """
    Sequential Tree-of-Thoughts BFS.

    - Branching factor controls how many children per node
    - Processes thoughts one at a time (no threading overhead)
    - Gurobi can use all CPU cores for each optimization

    Returns a list of (state, value) pairs.
    """
    tot_start_time = time.time()
    logger.info(f"ToT BFS starting | depth={depth} branching={branching} n_generate={n_generate} n_evaluate={n_evaluate}")

    task = SeatHarmonyTask()
    # Reset cache statistics for this ToT run
    task.cache_hits = 0
    task.cache_misses = 0
    root = task.get_initial_state(instance)
    logger.debug(f"Initial state created | guests={len(root.guests)} tables={len(root.venue.tables)}")

    frontier: List[SeatHarmonyState] = [root]
    scored_states: List[Tuple[SeatHarmonyState, float]] = []

    for level in range(depth):
        level_start_time = time.time()
        logger.info(f"ToT Level {level + 1}/{depth} starting | frontier_size={len(frontier)}")

        level_scored: List[Tuple[SeatHarmonyState, float]] = []

        # Collect all (state, thought) pairs to process at this level
        work_items: List[Tuple[SeatHarmonyState, str]] = []
        for state in frontier:
            thoughts = task.generate_thoughts(state, n_generate)[:branching]
            for thought in thoughts:
                work_items.append((state, thought))

        logger.info(f"ToT Level {level + 1}/{depth} | Generated {len(work_items)} work items (thoughts)")

        completed_count = 0
        failed_count = 0

        # Process work items sequentially
        for idx, (state, thought) in enumerate(work_items):
            thought_start_time = time.time()
            logger.info(f"ToT Processing thought {idx + 1}/{len(work_items)}: '{thought}'")

            try:
                child_state = task.apply_thought(state, thought)

                if child_state.layout is not None:
                    # Evaluate this child
                    evaluated = task.evaluate_states([child_state], n_evaluate)
                    level_scored.extend(evaluated)
                    completed_count += 1

                    thought_duration_ms = (time.time() - thought_start_time) * 1000
                    score = child_state.layout.score if child_state.layout else 0.0
                    logger.debug(f"Completed thought {idx + 1}/{len(work_items)}: {thought} | score={score:.2f} | {thought_duration_ms:.0f}ms")
                else:
                    failed_count += 1
                    logger.warning(f"Thought produced no layout: {thought}")

            except Exception as e:
                failed_count += 1
                logger.warning(f"Failed to apply thought: {thought} | error={type(e).__name__}: {e}")
                continue

        level_duration_ms = (time.time() - level_start_time) * 1000
        total_requests = task.cache_hits + task.cache_misses
        cache_hit_rate = (task.cache_hits / total_requests * 100) if total_requests > 0 else 0
        logger.info(f"ToT Level {level + 1}/{depth} completed | success={completed_count} failed={failed_count} | cache_hits={task.cache_hits} optimizations={task.cache_misses} (hit_rate={cache_hit_rate:.0f}%) | {level_duration_ms:.0f}ms")

        # Add this level's results to overall scored states
        scored_states.extend(level_scored)

        # Select top states for next level's frontier
        level_sorted = sorted(level_scored, key=lambda x: x[1], reverse=True)
        frontier = [s for s, _ in level_sorted[:branching]]

        if level_sorted:
            top_scores = [f"{s[1]:.2f}" for s in level_sorted[:3]]
            logger.debug(f"ToT Level {level + 1}/{depth} | Top scores: {', '.join(top_scores)}")

        # Early termination if no valid states
        if not frontier:
            logger.warning(f"ToT early termination at level {level + 1} - no valid states in frontier")
            break

    tot_duration_ms = (time.time() - tot_start_time) * 1000
    total_requests = task.cache_hits + task.cache_misses
    cache_hit_rate = (task.cache_hits / total_requests * 100) if total_requests > 0 else 0
    logger.info(f"ToT BFS completed | total_states={len(scored_states)} | cache_hits={task.cache_hits} optimizations={task.cache_misses} (hit_rate={cache_hit_rate:.0f}%) | {tot_duration_ms:.0f}ms")

    return scored_states


def _sequential_tot_bfs_generator(
    instance: Dict[str, Any],
    depth: int,
    branching: int = 3,
    n_generate: int = 3,
    n_evaluate: int = 3,
):
    """
    Generator version of Sequential Tree-of-Thoughts BFS.
    Yields JSON strings with progress updates.
    Finally yields the result as a JSON string.
    """
    tot_start_time = time.time()
    logger.info(f"ToT BFS Generator starting | depth={depth} branching={branching}")

    task = SeatHarmonyTask()
    # Reset cache statistics for this ToT run
    task.cache_hits = 0
    task.cache_misses = 0
    root = task.get_initial_state(instance)
    
    frontier: List[SeatHarmonyState] = [root]
    scored_states: List[Tuple[SeatHarmonyState, float]] = []

    # Calculate total optimizations for progress tracking
    # depth=1: branching optimizations (3)
    # depth=2: branching + branching*branching optimizations (3 + 9 = 12)
    total_optimizations = sum(branching ** (d + 1) for d in range(depth))
    current_optimization = 0

    # Initial progress
    yield json.dumps({
        "type": "progress",
        "percent": 0,
        "message": "Initializing optimization...",
        "currentStep": 0,
        "totalSteps": total_optimizations,
    }) + "\n"

    for level in range(depth):
        level_start_time = time.time()
        yield json.dumps({
            "type": "progress",
            "percent": int((current_optimization / total_optimizations) * 100),
            "message": f"Starting Level {level + 1} of {depth}...",
            "currentStep": current_optimization,
            "totalSteps": total_optimizations,
        }) + "\n"

        level_scored: List[Tuple[SeatHarmonyState, float]] = []

        # Collect work items
        work_items: List[Tuple[SeatHarmonyState, str]] = []
        for state in frontier:
            thoughts = task.generate_thoughts(state, n_generate)[:branching]
            for thought in thoughts:
                work_items.append((state, thought))

        total_work = len(work_items)
        if total_work == 0:
            break

        completed_count = 0

        # Process work items
        for idx, (state, thought) in enumerate(work_items):
            current_optimization += 1
            # Calculate progress as a percentage of total optimizations
            current_percent = min(99, int((current_optimization / total_optimizations) * 100))

            yield json.dumps({
                "type": "progress",
                "percent": current_percent,
                "message": f"Analyzing strategy {current_optimization} of {total_optimizations}...",
                "strategy": thought,
                "currentStep": current_optimization,
                "totalSteps": total_optimizations,
            }) + "\n"

            try:
                child_state = task.apply_thought(state, thought)

                if child_state.layout is not None:
                    # Evaluate
                    evaluated = task.evaluate_states([child_state], n_evaluate)
                    level_scored.extend(evaluated)
                    completed_count += 1
            except Exception as e:
                logger.warning(f"Failed to apply thought: {thought} | {e}")
                continue

        # Add this level's results
        scored_states.extend(level_scored)

        # Select top states for next level
        level_sorted = sorted(level_scored, key=lambda x: x[1], reverse=True)
        frontier = [s for s, _ in level_sorted[:branching]]

        if not frontier:
            break

    # Final processing - deduplicate by metrics similarity, keep top 3 by score
    # Two solutions with the same family/social/mixing percentages are duplicates
    # even if their exact assignments differ slightly
    unique_layouts = []
    seen_metrics = set()

    # Sort by score (value) descending to ensure highest scores first
    for state, value in sorted(scored_states, key=lambda x: x[1], reverse=True):
        if state.layout is None:
            continue
        layout_dict = layout_to_dict(state.layout)

        # Deduplicate by metrics (rounded to nearest 1%)
        # This ensures solutions with same objective percentages are considered duplicates
        breakdown = layout_dict.get("objective_breakdown", {})
        metrics_key = (
            round(breakdown.get("family_cohesion", 0)),
            round(breakdown.get("social_group_cohesion", 0)),
            round(breakdown.get("side_mixing", 0)),
        )

        if metrics_key in seen_metrics:
            continue  # Skip solutions with same metrics
        seen_metrics.add(metrics_key)

        unique_layouts.append({
            "value": value,
            "weights": state.weights,
            "notes": state.notes,
            "layout": layout_dict,
        })

        # Return top 3 unique layouts with highest scores
        if len(unique_layouts) >= 3:
            break

    yield json.dumps({
        "type": "result",
        "layouts": unique_layouts
    }) + "\n"


@app.post("/api/layouts/stream-generate")
def stream_generate_layouts(req: LayoutRequest) -> StreamingResponse:
    instance: Dict[str, Any] = {
        "guests": [g.dict() for g in req.guests],
        "tables": [t.dict() for t in req.tables],
        "settings": req.settings,
    }
    
    return StreamingResponse(
        _sequential_tot_bfs_generator(
            instance=instance,
            depth=req.tot.depth,
            branching=req.tot.branching,
            n_generate=req.tot.n_generate,
            n_evaluate=req.tot.n_evaluate,
        ),
        media_type="application/x-ndjson"
    )


@app.post("/api/layouts/generate")
def generate_layouts(req: LayoutRequest) -> Dict[str, Any]:
    # Set unique request ID for this request
    req_id = set_request_id()
    request_start_time = time.time()

    logger.info("=" * 50)
    logger.info(f"POST /api/layouts/generate | guests={len(req.guests)} tables={len(req.tables)}")
    logger.info(f"ToT params: depth={req.tot.depth} branching={req.tot.branching} n_generate={req.tot.n_generate} top_k={req.tot.top_k}")

    instance: Dict[str, Any] = {
        "guests": [g.dict() for g in req.guests],
        "tables": [t.dict() for t in req.tables],
        "settings": req.settings,
    }

    scored_states = _sequential_tot_bfs(
        instance=instance,
        depth=req.tot.depth,
        branching=req.tot.branching,
        n_generate=req.tot.n_generate,
        n_evaluate=req.tot.n_evaluate,
    )

    # Sort by value (score) descending and take top_k distinct layouts
    # Deduplicate by metrics similarity - solutions with same family/social/mixing
    # percentages are considered duplicates even if exact assignments differ
    unique_layouts: List[Dict[str, Any]] = []
    seen_metrics = set()
    duplicates_skipped = 0

    # Sort by score (value) descending - highest scores first
    for state, value in sorted(scored_states, key=lambda x: x[1], reverse=True):
        if state.layout is None:
            continue
        layout_dict = layout_to_dict(state.layout)

        # Deduplicate by metrics (rounded to nearest 1%)
        breakdown = layout_dict.get("objective_breakdown", {})
        metrics_key = (
            round(breakdown.get("family_cohesion", 0)),
            round(breakdown.get("social_group_cohesion", 0)),
            round(breakdown.get("side_mixing", 0)),
        )

        if metrics_key in seen_metrics:
            duplicates_skipped += 1
            continue  # Skip solutions with same metrics
        seen_metrics.add(metrics_key)

        unique_layouts.append(
            {
                "value": value,
                "weights": state.weights,
                "notes": state.notes,
                "layout": layout_dict,
            }
        )

        # Return top_k unique layouts with highest scores
        if len(unique_layouts) >= req.tot.top_k:
            break

    request_duration_ms = (time.time() - request_start_time) * 1000
    logger.info(f"Response: {len(unique_layouts)} unique layouts (skipped {duplicates_skipped} duplicates)")

    if unique_layouts:
        scores = [f"{l['value']:.2f}" for l in unique_layouts[:3]]
        notes = [l['notes'] for l in unique_layouts[:3]]
        logger.info(f"Top layouts: scores={scores} strategies={notes}")

    logger.info(f"POST /api/layouts/generate completed | {request_duration_ms:.0f}ms")
    logger.info("=" * 50)

    return {"layouts": unique_layouts}


@app.post("/api/layouts/explain")
def explain_layout(req: ExplainRequest) -> Dict[str, Any]:
    """
    Very simple textual explanation based on the numeric score and objective breakdown.
    This is a placeholder for a richer Gemini-backed explanation.
    """
    layout = req.layout
    score = layout.get("score", 0.0)
    breakdown = layout.get("objective_breakdown", {})

    parts: List[str] = []
    fc = breakdown.get("family_cohesion", 0.0)
    if fc:
        parts.append(
            f"Emphasizes keeping family members together (family cohesion weight {fc:.2f})."
        )
    sgc = breakdown.get("social_group_cohesion", 0.0)
    if sgc:
        parts.append(
            f"Keeps social groups together (social group cohesion weight {sgc:.2f})."
        )
    sm = breakdown.get("side_mixing", 0.0)
    if sm:
        parts.append(
            f"Encourages mixing between groom's and bride's sides (side mixing weight {sm:.2f})."
        )
    rp = breakdown.get("relationship_priority", 0.0)
    if rp:
        parts.append(
            f"Prioritizes closer relationships for better table assignments (relationship priority weight {rp:.2f})."
        )

    if not parts:
        parts.append(
            "Uses a neutral objective; primarily ensures everyone is seated within table capacities."
        )

    explanation = (
        f"This layout has an overall objective score of {score:.2f}. "
        + " ".join(parts)
    )

    return {"explanation": explanation}


# Category sets and helpers imported from centralized constants
from .constants import (
    STRATEGY_DESCRIPTIONS,
    get_guest_side as _get_guest_side,
)


def _get_strategy_description(notes: str) -> str:
    """Convert strategy code to human-readable description."""
    # Handle various formats: "boost_family", "Strategy: boost_family", etc.
    strategy_key = notes.lower().replace("strategy:", "").strip()
    return STRATEGY_DESCRIPTIONS.get(strategy_key, "optimized arrangement")


def _generate_fallback_explanations(table_guests: List[Dict[str, Any]]) -> Dict[str, str]:
    """
    Generate fallback explanations for guests when LLM call fails or returns incomplete results.

    Args:
        table_guests: List of guest dicts with 'id', 'name', and optionally 'group_id'

    Returns:
        Dict mapping guest_id -> explanation string
    """
    explanations = {}
    for g in table_guests:
        guest_name = g["name"]
        category = g.get("group_id") or "Uncategorized"

        if "Family" in category:
            explanations[g["id"]] = f"{guest_name} sits with family members as part of the seating arrangement."
        else:
            explanations[g["id"]] = f"{guest_name} is seated here as part of the optimized arrangement."

    return explanations


def _explain_guests_batch(
    table_guests: List[Dict[str, Any]],
    table: Dict[str, Any],
    table_index: int,
    all_tables: List[Dict[str, Any]],
    all_guests: List[Dict[str, Any]],
    assignments: Dict[str, str],
    weights: Dict[str, float],
    notes: str,
) -> Dict[str, str]:
    """
    Generate explanations for all guests at a table in a single LLM call.
    Returns a dict mapping guest_id -> explanation.
    """
    table_name = table.get("name", f"Table {table_index + 1}")
    logger.debug(f"Generating explanations for {table_name} | guests={len(table_guests)}")
    batch_start_time = time.time()

    # Get human-readable strategy description
    strategy_desc = _get_strategy_description(notes)

    # Build table context
    table_guest_names = [g["name"] for g in table_guests]
    table_categories = {}
    table_sides = {"groom's side": 0, "bride's side": 0, "neutral": 0}

    for g in table_guests:
        cat = g.get("group_id") or "Uncategorized"
        table_categories[cat] = table_categories.get(cat, 0) + 1
        side = _get_guest_side(cat)
        table_sides[side] += 1

    # Build guest details with richer context
    guest_details = []
    table_guest_ids = {g["id"] for g in table_guests}

    # Find where other guests from same categories are seated (for context)
    category_distribution = {}
    for g in all_guests:
        cat = g.get("group_id")
        if cat:
            assigned_table = assignments.get(g["id"])
            if cat not in category_distribution:
                category_distribution[cat] = {}
            if assigned_table:
                category_distribution[cat][assigned_table] = category_distribution[cat].get(assigned_table, 0) + 1

    for g in table_guests:
        category = g.get("group_id") or "Uncategorized"
        side = _get_guest_side(category)

        # Find other guests at this table with same category
        same_category_at_table = [
            other["name"] for other in table_guests
            if other["id"] != g["id"] and other.get("group_id") == category and category != "Uncategorized"
        ]

        # Check if category is split across tables
        cat_tables = category_distribution.get(category, {})
        is_category_split = len(cat_tables) > 1

        # Build context string for this guest
        context_parts = []
        if same_category_at_table:
            context_parts.append(f"seated with {len(same_category_at_table)} others from same group")
        if is_category_split and category != "Uncategorized":
            other_tables = [t for t in cat_tables.keys() if t != table.get("id")]
            if other_tables:
                context_parts.append(f"some {category} guests are at other tables")
        if g.get("importance", 0) > 0:
            context_parts.append("VIP guest")

        guest_details.append({
            "name": g["name"],
            "category": category,
            "side": side,
            "context": "; ".join(context_parts) if context_parts else "flexible placement",
        })
    
    # Build natural context about the table
    table_summary = []
    if len(table_categories) == 1:
        table_summary.append(f"all guests are from {list(table_categories.keys())[0]}")
    else:
        main_category = max(table_categories.items(), key=lambda x: x[1])[0]
        table_summary.append(f"mostly {main_category} with some mixing")

    # Add side mixing info
    groom_count = table_sides.get("groom's side", 0)
    bride_count = table_sides.get("bride's side", 0)
    if groom_count > 0 and bride_count > 0:
        table_summary.append(f"mixed table with {groom_count} from groom's side and {bride_count} from bride's side")
    elif groom_count > 0:
        table_summary.append("primarily groom's side guests")
    elif bride_count > 0:
        table_summary.append("primarily bride's side guests")

    # Format guest list with side information
    guest_list = "\n".join([
        f"- {g['name']} ({g['category']}, {g['side']}) — {g['context']}"
        for g in guest_details
    ])

    # Describe the weights in natural language
    family_weight = weights.get("family_cohesion", 0.5)
    social_weight = weights.get("social_group_cohesion", 0.5)
    mixing_weight = weights.get("side_mixing", 0.5)

    weight_priorities = []
    if family_weight >= 0.7:
        weight_priorities.append("keeping families together was a high priority")
    if social_weight >= 0.7:
        weight_priorities.append("keeping friend groups together was important")
    if mixing_weight >= 0.7:
        weight_priorities.append("mixing bride and groom sides was encouraged")
    if not weight_priorities:
        weight_priorities.append("a balanced approach was used")

    prompt = f"""You are explaining wedding seating decisions to the couple. For each guest, provide ONE natural, concise sentence explaining why they are seated at this table.

SEATING STRATEGY:
This arrangement was created {strategy_desc}. Specifically, {'; '.join(weight_priorities)}.

TABLE CONTEXT:
{table_name} has {len(table_guests)} guests. {'. '.join(table_summary)}.

GUESTS AT THIS TABLE:
{guest_list}

INSTRUCTIONS:
1. Write in THIRD PERSON (e.g., "Sarah sits with..." NOT "You sit...")
2. ONE sentence per guest — be concise but meaningful
3. DO NOT state the table name or number — the user already knows where they sit
4. Focus on WHY they are seated here:
   - Seated with their group: "joins other [category] guests at this table"
   - Group was split due to table size: "seated here while other [group] members are nearby"
   - Mixed table for mingling: "seated here to bring together both sides of the family"
   - VIP guest: mention their importance naturally
   - Flexible placement: "seated here to balance the table"
5. Be warm and natural — this is a wedding, not a corporate event
6. NEVER use technical terms like "optimization", "algorithm", "constraints", or "cohesion"

OUTPUT FORMAT:
Guest: [Full Name]
Explanation: [ONE sentence]

EXAMPLES:
Guest: Sarah Cohen
Explanation: Sarah joins her family members from the groom's side at this table.

Guest: David Miller
Explanation: David is seated here alongside his college friends to keep the group together.

Guest: Emma Thompson
Explanation: Emma brings a friendly connection between the bride's and groom's sides at this mixed table.

Guest: Michael Chen
Explanation: Michael is placed here to round out the table with good company.

Now write ONE natural sentence for each of the {len(table_guests)} guests:"""

    try:
        # Reduced max_tokens since we're generating one sentence per guest
        # Lower temperature (0.4) for more consistent, factual explanations
        logger.debug(f"Calling LLM for {table_name} explanations | model=llama-3.3-70b-versatile (Groq) temp=0.4")
        llm_start_time = time.time()
        
        # Estimate tokens needed (rough estimate: ~1.3 tokens per character)
        estimated_tokens = int(len(prompt) * 1.3) + 400  # prompt + max response tokens
        
        # Wait if needed before making request (proactive rate limiting)
        _groq_rate_limiter.wait_if_needed(estimated_tokens)
        
        # Retry logic with exponential backoff for rate limits
        max_retries = 5  # Increased retries
        base_delay = 2  # seconds
        response = None
        actual_tokens = 0
        for attempt in range(max_retries):
            try:
                # Use direct Groq API call to get actual token usage
                response, actual_tokens = _call_groq_with_usage(
                    prompt=prompt,
                    model="llama-3.3-70b-versatile",
                    temperature=0.4,
                    max_tokens=400
                )
                llm_duration_ms = (time.time() - llm_start_time) * 1000
                
                # Record request and actual token usage (tracks all limits: RPM, RPD, TPM, TPD)
                _groq_rate_limiter.record_request(actual_tokens)
                
                logger.debug(f"LLM response received | {llm_duration_ms:.0f}ms | response_len={len(response)} | tokens_used={actual_tokens}")
                break  # Success, exit retry loop
            except requests.HTTPError as e:
                # Handle 429 rate limit errors specifically
                if e.response is not None and e.response.status_code == 429:
                    error_str = str(e)
                    
                    # Priority 1: Use reset time from headers (most accurate)
                    headers = _groq_rate_limiter.last_rate_limit_headers
                    wait_time = None
                    
                    # Check if we have token reset time (most common limit hit)
                    if headers.get('reset_tokens_seconds') is not None:
                        wait_time = headers['reset_tokens_seconds']
                        # Add small buffer to ensure tokens are available
                        wait_time = max(wait_time + 0.1, 0.1)
                    # Fallback to request reset time
                    elif headers.get('reset_requests_seconds') is not None:
                        wait_time = headers['reset_requests_seconds']
                        wait_time = max(wait_time + 0.1, 0.1)
                    
                    # Priority 2: Try to parse wait time from error message
                    if wait_time is None:
                        wait_time = _groq_rate_limiter.parse_rate_limit_error(error_str)
                    
                    # Priority 3: Use exponential backoff as last resort
                    if wait_time is None:
                        wait_time = base_delay * (2 ** attempt)
                    
                    if attempt < max_retries - 1:
                        status = _groq_rate_limiter.get_status()
                        logger.warning(
                            f"Rate limit hit for {table_name} | attempt {attempt + 1}/{max_retries} | "
                            f"waiting {wait_time:.2f}s | "
                            f"RPM: {status['rpm']['used']}/{status['rpm']['limit']} | "
                            f"RPD: {status['rpd']['used']}/{status['rpd']['limit']} | "
                            f"TPM: {status['tpm']['used']}/{status['tpm']['limit']} | "
                            f"TPD: {status['tpd']['used']}/{status['tpd']['limit']}"
                        )
                        time.sleep(wait_time)
                        # Clean old entries after waiting to free up capacity
                        _groq_rate_limiter._clean_old_entries(time.time())
                        continue
                    else:
                        logger.error(f"Rate limit exceeded after {max_retries} attempts for {table_name}")
                        raise  # Re-raise if all retries exhausted
                else:
                    # Non-429 HTTP errors - re-raise immediately
                    raise
            except Exception as e:
                error_str = str(e)
                # Check for other rate limit error patterns (from string representations)
                if "rate limit" in error_str.lower() or "429" in error_str or "RESOURCE_EXHAUSTED" in error_str or "TPM" in error_str:
                    # Priority 1: Use reset time from headers (most accurate)
                    headers = _groq_rate_limiter.last_rate_limit_headers
                    wait_time = None
                    
                    # Check if we have token reset time (most common limit hit)
                    if headers.get('reset_tokens_seconds') is not None:
                        wait_time = headers['reset_tokens_seconds']
                        # Add small buffer to ensure tokens are available
                        wait_time = max(wait_time + 0.1, 0.1)
                    # Fallback to request reset time
                    elif headers.get('reset_requests_seconds') is not None:
                        wait_time = headers['reset_requests_seconds']
                        wait_time = max(wait_time + 0.1, 0.1)
                    
                    # Priority 2: Try to parse wait time from error message
                    if wait_time is None:
                        wait_time = _groq_rate_limiter.parse_rate_limit_error(error_str)
                    
                    # Priority 3: Use exponential backoff as last resort
                    if wait_time is None:
                        wait_time = base_delay * (2 ** attempt)
                    
                    if attempt < max_retries - 1:
                        logger.warning(f"Rate limit error for {table_name} | attempt {attempt + 1}/{max_retries} | waiting {wait_time:.2f}s...")
                        time.sleep(wait_time)
                        _groq_rate_limiter._clean_old_entries(time.time())
                        continue
                    else:
                        logger.error(f"Rate limit exceeded after {max_retries} attempts for {table_name}")
                        raise
                else:
                    raise  # Re-raise non-rate-limit errors immediately
        
        if response is None:
            raise Exception("Failed to get LLM response after all retries")
        
        # Parse the response to extract individual explanations
        explanations = {}
        current_guest = None
        current_explanation = []
        
        lines = response.split('\n')
        for line in lines:
            line = line.strip()
            if not line:
                # Empty line - if we have a guest and explanation, save it
                if current_guest and current_explanation:
                    explanations[current_guest] = ' '.join(current_explanation).strip()
                    current_guest = None
                    current_explanation = []
                continue
                
            # Check for "Guest:" pattern (case insensitive, with or without colon)
            if line.lower().startswith('guest'):
                # Save previous guest if exists
                if current_guest and current_explanation:
                    explanations[current_guest] = ' '.join(current_explanation).strip()
                # Extract guest name
                parts = line.split(':', 1)
                if len(parts) > 1:
                    current_guest = parts[1].strip()
                else:
                    # Try to extract name after "Guest"
                    name_part = line.replace('Guest', '').strip()
                    if name_part:
                        current_guest = name_part
                current_explanation = []
            elif line.lower().startswith('explanation'):
                # Extract explanation text (should be one sentence)
                parts = line.split(':', 1)
                if len(parts) > 1:
                    explanation_text = parts[1].strip()
                    if explanation_text:
                        current_explanation.append(explanation_text)
            elif current_guest and line:
                # Continuation of explanation (shouldn't happen with one sentence, but handle it)
                if not line.lower().startswith(('guest', 'explanation', '---', '===')):
                    current_explanation.append(line)
        
        # Save last guest
        if current_guest and current_explanation:
            explanations[current_guest] = ' '.join(current_explanation).strip()
        
        # Map guest names back to IDs (fuzzy matching for robustness)
        name_to_id = {g["name"]: g["id"] for g in table_guests}
        # Also create lowercase mapping for case-insensitive matching
        name_to_id_lower = {g["name"].lower(): g["id"] for g in table_guests}
        
        result = {}
        for name, explanation in explanations.items():
            # Try exact match first
            guest_id = name_to_id.get(name)
            if not guest_id:
                # Try case-insensitive match
                guest_id = name_to_id_lower.get(name.lower())
            if guest_id and explanation:
                result[guest_id] = explanation
        
        # If parsing failed or incomplete, provide fallback explanations for missing guests
        missing_guests = [g for g in table_guests if g["id"] not in result]
        if missing_guests:
            fallback = _generate_fallback_explanations(missing_guests)
            result.update(fallback)
        fallback_count = len(missing_guests)

        batch_duration_ms = (time.time() - batch_start_time) * 1000
        logger.debug(f"{table_name} explanations complete | parsed={len(result) - fallback_count} fallback={fallback_count} | {batch_duration_ms:.0f}ms")

        return result
        
    except Exception as e:
        # Fallback if LLM call fails
        batch_duration_ms = (time.time() - batch_start_time) * 1000
        logger.error(f"LLM explanation failed for {table_name} | error={type(e).__name__}: {e} | {batch_duration_ms:.0f}ms")
        logger.info(f"Using fallback explanations for {table_name}")
        return _generate_fallback_explanations(table_guests)


@app.post("/api/layouts/explain-guests")
def explain_guests_seating(req: ExplainGuestsRequest) -> Dict[str, Any]:
    """
    Generate explanations for all guests, batched by table.
    Returns a dict mapping guest_id -> explanation.
    """
    req_id = set_request_id()
    request_start_time = time.time()

    logger.info("=" * 50)
    logger.info(f"POST /api/layouts/explain-guests | guests={len(req.guests)} tables={len(req.tables)}")
    logger.info(f"Strategy: {req.notes}")

    assignments = req.layout.get("assignments", {})
    all_guests_dict = {g.id: g.dict() for g in req.guests}
    all_tables_dict = {t.id: t.dict() for t in req.tables}

    # Group guests by table using helper function
    table_to_guests = _group_guests_by_table(assignments, all_guests_dict)
    logger.debug(f"Guests grouped into {len(table_to_guests)} tables")

    # Generate explanations for each table (batched)
    all_explanations: Dict[str, str] = {}
    tables_list = list(req.tables)

    for table_idx, (table_id, table_guests) in enumerate(table_to_guests.items()):
        if not table_guests or table_id not in all_tables_dict:
            continue

        table = all_tables_dict[table_id]
        table_index = next((i for i, t in enumerate(tables_list) if t.id == table_id), 0)

        logger.debug(f"Processing table {table_idx + 1}/{len(table_to_guests)} | {table.get('name', table_id)} | {len(table_guests)} guests")

        # Add delay between requests to avoid rate limits (except for first table)
        if table_idx > 0:
            # Check available tokens and wait if needed before processing next table
            estimated_next_tokens = 1000  # Conservative estimate for next request
            wait_time = _groq_rate_limiter.wait_if_needed(estimated_next_tokens)
            if wait_time == 0:
                # If no wait needed, still add a small delay to be safe
                time.sleep(1.0)  # Increased from 0.5s to 1.0s delay between table requests
            status = _groq_rate_limiter.get_status()
            logger.debug(f"Rate limiter status | RPM: {status['rpm']['used']}/{status['rpm']['limit']} | RPD: {status['rpd']['used']}/{status['rpd']['limit']} | TPM: {status['tpm']['used']}/{status['tpm']['limit']} | TPD: {status['tpd']['used']}/{status['tpd']['limit']}")

        # Generate batch explanation for this table
        table_explanations = _explain_guests_batch(
            table_guests=table_guests,
            table=table,
            table_index=table_index,
            all_tables=[t.dict() for t in req.tables],
            all_guests=[g.dict() for g in req.guests],
            assignments=assignments,
            weights=req.weights,
            notes=req.notes,
        )

        all_explanations.update(table_explanations)

    request_duration_ms = (time.time() - request_start_time) * 1000
    logger.info(f"Generated {len(all_explanations)} guest explanations | {request_duration_ms:.0f}ms")
    logger.info("=" * 50)

    return {"explanations": all_explanations}


@app.post("/api/export/excel")
def export_excel(req: ExcelExportRequest):
    """
    Generate and return an Excel file with seating plan data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    logger.info(f"POST /api/export/excel | guests={len(req.guests)} tables={len(req.tables)}")

    # Create workbook
    wb = Workbook()

    # Get layout assignments
    assignments = req.layout.get("assignments", {})

    # Build guest and table lookups
    guests_dict = {g.id: g for g in req.guests}
    tables_dict = {t.id: t for t in req.tables}

    # Group guests by table using helper function
    table_to_guests = _group_guests_by_table(assignments, guests_dict)

    # ============ SHEET 1: Seating Plan ============
    ws1 = wb.active
    ws1.title = "Seating Plan"
    ws1.sheet_properties.tabColor = "8A8E75"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="8A8E75", end_color="8A8E75", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    # Headers
    include_dietary = req.options.get("include_dietary", True)
    
    headers1 = ["Guest Name", "Table Name", "Table #", "Group/Category"]
    if include_dietary:
        headers1.append("Dietary Restrictions")
    headers1.append("Notes")

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Data rows - sorted by table, then alphabetically by guest name
    row_num = 2
    sorted_tables = sorted(
        [(tid, tguests) for tid, tguests in table_to_guests.items() if isinstance(tguests, list)],
        key=lambda x: tables_dict.get(x[0]).name if x[0] in tables_dict else ""
    )

    for table_id, guests_list in sorted_tables:
        table = tables_dict.get(table_id)
        if not table:
            continue

        # Sort guests alphabetically within each table
        sorted_guests = sorted(guests_list, key=lambda g: g.name)

        for guest in sorted_guests:
            # Build row values dynamically
            row_values = [
                guest.name,
                table.name,
                int(table.name.replace("Table ", "")) if "Table " in table.name else row_num - 1,
                guest.group_id or "Uncategorized"
            ]
            
            if include_dietary:
                dietary = ", ".join([t for t in guest.tags if t.lower() in ["vegetarian", "vegan", "gluten-free", "kosher", "halal", "allergies"]])
                row_values.append(dietary)
            
            row_values.append("") # Notes
            
            # Write cells
            for col, val in enumerate(row_values, 1):
                ws1.cell(row=row_num, column=col, value=val).border = border

            # Alternating row color
            if row_num % 2 == 0:
                for col in range(1, len(row_values) + 1):
                    ws1.cell(row=row_num, column=col).fill = alt_fill

            row_num += 1

    # Auto-adjust column widths
    for col_idx, col in enumerate(ws1.columns, 1):
        column_letter = get_column_letter(col_idx)
        
        # Check header value to see if this is the Notes column
        header_val = col[0].value
        if header_val == "Notes":
            ws1.column_dimensions[column_letter].width = 60
            continue

        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws1.column_dimensions[column_letter].width = min(max_length + 2, 40)

    # ============ SHEET 2: Table Summary ============
    if req.options.get("include_table_details", True):
        ws2 = wb.create_sheet("Table Summary")
        ws2.sheet_properties.tabColor = "68604D"

        headers2 = ["Table Name", "Capacity", "Guests Seated", "Group Distribution", "Zone"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row_num = 2
        for table in sorted(req.tables, key=lambda t: t.name):
            guests_at_table = table_to_guests.get(table.id, [])
            if not isinstance(guests_at_table, list):
                guests_at_table = []

            # Group distribution
            group_counts: Dict[str, int] = {}
            for g in guests_at_table:
                group = g.group_id or "Uncategorized"
                group_counts[group] = group_counts.get(group, 0) + 1

            group_dist = ", ".join([f"{k}: {v}" for k, v in sorted(group_counts.items())])

            ws2.cell(row=row_num, column=1, value=table.name).border = border
            ws2.cell(row=row_num, column=2, value=table.capacity).border = border
            ws2.cell(row=row_num, column=3, value=len(guests_at_table)).border = border
            ws2.cell(row=row_num, column=4, value=group_dist).border = border
            ws2.cell(row=row_num, column=5, value=table.zone or "").border = border

            if row_num % 2 == 0:
                for col in range(1, 6):
                    ws2.cell(row=row_num, column=col).fill = alt_fill

            row_num += 1

        # Auto-adjust column widths
        for col_idx in range(1, 6):
            max_length = len(headers2[col_idx - 1])
            column_letter = get_column_letter(col_idx)
            for row in range(2, row_num):
                cell_value = ws2.cell(row=row, column=col_idx).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            ws2.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # ============ SHEET 3: Group Analysis ============
    if req.options.get("include_vendor_summary", False): # Reusing this flag for Group Analysis
        ws3 = wb.create_sheet("Group Analysis")
        ws3.sheet_properties.tabColor = "D5C7AD"

        headers3 = ["Group Name", "Guest Name", "Assigned Table", "Table #"]
        for col, header in enumerate(headers3, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Group guests by their category
        guests_by_group: Dict[str, List[GuestIn]] = {}
        for guest in req.guests:
            group = guest.group_id or "Uncategorized"
            if group not in guests_by_group:
                guests_by_group[group] = []
            guests_by_group[group].append(guest)

        row_num = 2
        for group_idx, group in enumerate(sorted(guests_by_group.keys())):
            # Determine background color for this entire group
            # Use alt_fill for odd groups, white for even groups
            is_alt_group = group_idx % 2 != 0
            group_fill = alt_fill if is_alt_group else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Sorting guests within group
            group_guests = sorted(guests_by_group[group], key=lambda g: g.name)

            start_row = row_num
            for i, guest in enumerate(group_guests):
                table_id = assignments.get(guest.id)
                table = tables_dict.get(table_id)
                table_name = table.name if table else "Unseated"
                
                # Parse table number
                table_num_val = ""
                if table and "Table " in table.name:
                    try:
                        table_num_val = int(table.name.replace("Table ", ""))
                    except:
                        pass
                
                # Column 1: Group Name
                group_val = group if i == 0 else ""
                c1 = ws3.cell(row=row_num, column=1, value=group_val)
                c1.border = border
                c1.fill = group_fill # Apply uniform group fill
                
                if i == 0:
                     c1.font = Font(bold=True)
                     c1.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                c2 = ws3.cell(row=row_num, column=2, value=guest.name)
                c2.border = border
                c2.fill = group_fill

                c3 = ws3.cell(row=row_num, column=3, value=table_name)
                c3.border = border
                c3.fill = group_fill

                c4 = ws3.cell(row=row_num, column=4, value=table_num_val)
                c4.border = border
                c4.fill = group_fill

                row_num += 1
            
            # Merge Group Name cells
            end_row = row_num - 1
            if end_row > start_row:
                ws3.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            
            # Add a separator row between groups (optional, or just logic as above is fine)

        # Auto-adjust column widths
        for col_idx in range(1, 5):
            max_length = len(headers3[col_idx - 1])
            column_letter = get_column_letter(col_idx)
            for row in range(2, row_num):
                cell_value = ws3.cell(row=row, column=col_idx).value
                if cell_value and len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            ws3.column_dimensions[column_letter].width = min(max_length + 2, 40)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    logger.info(f"Excel export generated successfully | sheets={len(wb.worksheets)}")

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=SeatHarmony_SeatingPlan.xlsx"}
    )

